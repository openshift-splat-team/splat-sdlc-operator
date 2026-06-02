#!/usr/bin/env python3
"""
Seed the Jira simulator from Jira xlsx exports.

Reads every .xlsx file in test_data/ (or paths supplied as arguments),
extracts issues from the "Work Item Fields" sheet, and bulk-imports them
into the simulator via POST /api/admin/import.

Issues are imported with their original Jira keys (e.g. SPLAT-2724) so
they can be referenced by key just like in real Jira.  Already-existing
issues are skipped (idempotent).

Comments from the "Work Item Activity" sheet (Type=COMMENT rows) are
imported as simulator comments on the relevant issue.

Usage:
    uv run python scripts/jira_seed.py [--url URL] [file ...]

    --url   Simulator base URL (default: http://localhost:8080)
    file    One or more xlsx files or glob patterns
            (default: test_data/*.xlsx)
"""
from __future__ import annotations

import argparse
import glob
import json
import sys
import urllib.error
import urllib.request
import warnings
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "openpyxl is required: pip install openpyxl  or  uv pip install openpyxl"
    )

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ── Column name → simulator field mapping ─────────────────────────────────────

# Each entry is (xlsx_column_name, simulator_field).
# The first match wins when multiple column names map to the same field.
FIELD_MAP: list[tuple[str, str]] = [
    ("Issue key",                              "key"),
    ("Summary",                                "summary"),
    ("Issue Type",                             "issuetype"),
    ("Status",                                 "status"),
    ("Resolution",                             "resolution"),
    ("Description",                            "description"),
    ("Custom field (Acceptance Criteria)",     "acceptance_criteria"),
    ("Custom field (Story Points)",            "story_points"),
    ("Custom field (Story point estimate)",    "story_points"),   # alternate label
    ("Custom field (Story point)",             "story_points"),   # alternate label
    ("Custom field (Parent Link)",             "epic_link_key"),  # epic's parent in some exports
    ("Parent key",                             "parent_key"),
    # Labels are handled separately — Jira exports use one column per label,
    # all named "Labels", so they're collected via _collect_labels().
]

# Jira status → simulator status
STATUS_MAP: dict[str, str] = {
    "new":                    "To Do",
    "open":                   "To Do",
    "backlog":                "To Do",
    "refinement":             "To Do",
    "selected for development": "To Do",
    "todo":                   "To Do",
    "to do":                  "To Do",
    "in progress":            "In Progress",
    "in review":              "In Progress",
    "review":                 "In Progress",
    "done":                   "Done",
    "closed":                 "Done",
    "resolved":               "Done",
    "won't fix":              "Won't Do",
    "won't do":               "Won't Do",
    "wontfix":                "Won't Do",
}


def _normalise_status(raw: str | None) -> str:
    if not raw:
        return "To Do"
    return STATUS_MAP.get(raw.strip().lower(), raw.strip())


def _str(val: object) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _float(val: object) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── xlsx parsing ──────────────────────────────────────────────────────────────

def parse_fields_sheet(ws) -> list[dict]:
    """Return a list of issue dicts from the Work Item Fields sheet."""
    headers = [cell.value for cell in ws[1]]
    col_index: dict[str, int] = {h: i for i, h in enumerate(headers) if h}

    # Build a lookup: simulator_field → column index (first match wins)
    field_to_col: dict[str, int] = {}
    for col_name, field in FIELD_MAP:
        if col_name in col_index and field not in field_to_col:
            field_to_col[field] = col_index[col_name]

    # Jira exports one column per label, all named "Labels" — collect all their indices.
    label_col_indices: list[int] = [i for i, h in enumerate(headers) if h == "Labels"]

    issues = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        def get(field: str) -> object:
            idx = field_to_col.get(field)
            return row[idx] if idx is not None else None

        key = _str(get("key"))
        if not key:
            continue

        # Collect all non-empty values from every "Labels" column.
        labels = [str(row[i]) for i in label_col_indices if row[i] is not None and str(row[i]).strip()]

        summary = _str(get("summary")) or ""
        description = _str(get("description")) or ""
        acceptance = _str(get("acceptance_criteria"))
        if acceptance:
            sep = "\n\n" if description else ""
            description = f"{description}{sep}h3. Acceptance Criteria\n{acceptance}"

        # For epics, "parent_key" in the xlsx is actually the strategic epic above
        # (e.g. OCPSTRAT-2933). We store it as parent_key so it's visible in the UI.
        parent_key = _str(get("parent_key"))
        epic_link_key = _str(get("epic_link_key"))

        # Avoid self-referential or circular parent links from the export
        if parent_key == key:
            parent_key = None
        if epic_link_key == key:
            epic_link_key = None

        issues.append({
            "key": key,
            "summary": summary,
            "description": description or None,
            "issuetype": _str(get("issuetype")) or "Story",
            "status": _normalise_status(_str(get("status"))),
            "resolution": _str(get("resolution")),
            "story_points": _float(get("story_points")),
            "epic_link_key": epic_link_key,
            "parent_key": parent_key,
            "labels": labels,
        })

    return issues


def parse_activity_sheet(ws) -> dict[str, list[str]]:
    """
    Return {issue_key: [comment_body, ...]} from the Work Item Activity sheet.
    Only rows with Type == "COMMENT" are included.
    """
    headers = [cell.value for cell in ws[1]]
    col_index = {h: i for i, h in enumerate(headers) if h}

    comments: dict[str, list[str]] = {}
    # Activity sheets are named "Work Item Activity {KEY}"
    # The issue key is embedded in the sheet title.
    title_parts = ws.title.split()
    issue_key = title_parts[-1].upper() if title_parts else ""

    type_col = col_index.get("Type")
    details_col = col_index.get("Details")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rtype = _str(row[type_col]) if type_col is not None else None
        if rtype and rtype.upper() == "COMMENT":
            body = _str(row[details_col]) if details_col is not None else None
            if body and issue_key:
                comments.setdefault(issue_key, []).append(body)

    return comments


def read_xlsx(path: str) -> tuple[list[dict], dict[str, list[str]]]:
    """Return (issues, comments_by_key) from a Jira export xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    issues: list[dict] = []
    comments: dict[str, list[str]] = {}

    for ws in wb.worksheets:
        title = ws.title
        if "Work Item Fields" in title or title == wb.sheetnames[0]:
            issues = parse_fields_sheet(ws)
        elif "Work Item Activity" in title:
            comments.update(parse_activity_sheet(ws))

    return issues, comments


# ── HTTP helpers ──────────────────────────────────────────────────────────────

def _post(url: str, payload: dict) -> dict:
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as resp:
            return json.load(resp)
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors="replace")
        raise SystemExit(f"HTTP {e.code} from {url}: {body}") from None


def _check_alive(base_url: str) -> None:
    try:
        with urllib.request.urlopen(f"{base_url}/rest/api/2/serverInfo") as r:
            info = json.load(r)
            print(f"  Connected to: {info.get('serverTitle', base_url)}")
    except Exception as exc:
        sys.exit(f"Cannot reach simulator at {base_url}: {exc}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--url", default="http://localhost:8080", help="Simulator base URL")
    parser.add_argument("--force", action="store_true", help="Re-import issues that already exist (overwrites fields including labels)")
    parser.add_argument("files", nargs="*", help="xlsx files or glob patterns (default: test_data/*.xlsx)")
    args = parser.parse_args()

    # Resolve file list
    patterns = args.files or ["test_data/*.xlsx"]
    paths: list[str] = []
    for pattern in patterns:
        expanded = glob.glob(pattern)
        if not expanded and Path(pattern).exists():
            expanded = [pattern]
        paths.extend(expanded)

    if not paths:
        sys.exit(f"No xlsx files found matching: {patterns}")

    print(f"\nJira simulator seed — {args.url}")
    _check_alive(args.url)
    print()

    all_issues: list[dict] = []
    all_comments: dict[str, list[str]] = {}

    for path in sorted(paths):
        issues, comments = read_xlsx(path)
        valid = [i for i in issues if i.get("key") and i.get("summary")]
        print(f"  {path}: {len(valid)} issue(s), {sum(len(v) for v in comments.values())} comment(s)")
        all_issues.extend(valid)
        for k, cs in comments.items():
            all_comments.setdefault(k, []).extend(cs)

    if not all_issues:
        sys.exit("No issues found in the xlsx file(s).")

    # Sort: epics first so parent references resolve correctly
    type_order = {"epic": 0, "story": 1, "subtask": 2}
    all_issues.sort(key=lambda i: type_order.get(i.get("issuetype", "").lower(), 1))

    # Deduplicate by key (keep last seen — last file wins)
    seen: dict[str, dict] = {}
    for issue in all_issues:
        seen[issue["key"]] = issue
    unique_issues = list(seen.values())

    print(f"\nImporting {len(unique_issues)} issue(s){' (--force: overwriting existing)' if args.force else ''}...")
    result = _post(f"{args.url}/api/admin/import", {"issues": unique_issues, "force": args.force})
    print(f"  Imported : {result['imported']}")
    print(f"  Skipped  : {result['skipped']} (already exist)")
    if result.get("skipped_keys"):
        print(f"  Keys skipped: {', '.join(result['skipped_keys'][:10])}", end="")
        if len(result["skipped_keys"]) > 10:
            print(f" ... (+{len(result['skipped_keys']) - 10} more)", end="")
        print()

    # Import comments for issues that were successfully imported
    if all_comments:
        imported_keys = {i["key"] for i in unique_issues} - set(result.get("skipped_keys", []))
        comment_count = 0
        for key, bodies in all_comments.items():
            if key not in imported_keys:
                continue
            for body in bodies:
                _post(f"{args.url}/rest/api/2/issue/{key}/comment", {"body": body})
                comment_count += 1
        if comment_count:
            print(f"  Comments : {comment_count} imported")

    print(f"\nBrowse: {args.url}/ui")


if __name__ == "__main__":
    main()
